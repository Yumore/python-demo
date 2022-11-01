#!/usr/bin/python3
# -*- coding: utf-8 -*-
import datetime
import hashlib
import os
import sys

from androguard.core.bytecodes.apk import APK
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def main(path):
    if os.path.isdir(path):
        wd = WriteData()
        wd.write_data(is_init=True)
        for root, dirs, files in os.walk(path):
            for f in range(len(files)):
                info = get_apk_info_f(root, files[f])
                print(info)
                wd.write_data(row=f, data=info)
    else:
        print('参数输入有误，不是一个目录...')


def get_apk_info_f(root, file):
    apk_path = os.path.join(root, file)
    apk_info = []
    try:
        android_guard = APK(apk_path)
        if android_guard.is_valid_APK():
            apk_info.append(file)
            apk_info.append(get_file_md5_f(apk_path))
            apk_info.append(get_cert_md5_f(android_guard))
            apk_info.append(android_guard.get_app_name())
            apk_info.append(android_guard.get_package())
            apk_info.append(android_guard.get_androidversion_name())
    except Exception as e:
        print(file + ' ->>', e)

    return apk_info


def get_cert_md5_f(android_guard):
    cert_md5 = ''
    certs = set(android_guard.get_certificates_der_v2() + [android_guard.get_certificate_der(x) for x in android_guard.get_signature_names()])
    for cert in certs:
        cert_md5 = hashlib.md5(cert).hexdigest()

    return cert_md5


def get_file_md5_f(file):
    with open(file, 'rb') as f:
        md5obj = hashlib.md5()
        md5obj.update(f.read())
        md5 = md5obj.hexdigest()
        md5 = str(md5).lower()

    return md5


class WriteData:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.today = str(datetime.date.today())
        self.font = Font(name='等线', size=11)
        self.fill = PatternFill("solid", fgColor="E0EEE0")
        self.title = ['filename', 'file_md5', 'cert_md5', 'app_name', 'pkg_name', 'app_version']

    def write_data(self, row=1, data=None, is_init=False):
        if not is_init:
            for col in range(1, len(data) + 1):
                operate = self.ws.cell(row=row + 2, column=col, value=data[col - 1])
                operate.font = self.font
        else:
            for i in range(1, len(self.title) + 1):
                operate = self.ws.cell(row=row, column=i, value=self.title[i - 1])
                operate.font = self.font
                operate.fill = self.fill

        self.wb.save(os.path.join(os.getcwd(), self.today + '.xlsx'))


if __name__ == '__main__':
    main(sys.argv[1])
